from os import getcwd
from datetime import date
from csv import reader, writer, register_dialect, QUOTE_NONE
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Alignment

def getCSVdata(file, dlm=",", enc="UTF-8"):
    """
    Gets data from a CSV file and returns it as a 2D list.
    :param file: str; CSV file to process
    :param dlm: str; delimiter, default ,
    :param enc: str; encoding, default UTF-8
    :return: list; CSV data in a 2D list
    """
    with open(file, encoding=enc) as stats:
        return list(reader(stats, delimiter=dlm))

# creates path
path = getcwd().replace("\\", "\\\\") + "\\\\"
# date variables
today = date.today()
prev_month_year = today.year
prev_month_month = today.month - 1
if prev_month_month == 0:
    prev_month_year -= 1
    prev_month_month = 12
prev_month_month = str(prev_month_month)
months = { 1: "Januar",
           2: "Februar",
           3: "Marec",
           4: "April",
           5: "Maj",
           6: "Junij",
           7: "Julij",
           8: "Avgust",
           9: "September",
          10: "Oktober",
          11: "November",
          12: "December"}
if len(prev_month_month) == 1: prev_month_month = "0" + prev_month_month
prev_month_year = str(prev_month_year)
# importing CSV data
register_dialect('myDialect', delimiter=';', quoting=QUOTE_NONE)
GAdW_prev = getCSVdata(path.replace("\\\\", "/") + "GAdW_Avtenta_CSV_{}-{}.csv".format(prev_month_year, prev_month_month), ";")
stats = [["nolabel", []]]
for element in GAdW_prev:
    for stat in sorted(stats):
        if element[0] == stat[0]:
            stat[1].append([float(item) for item in element[1:5]] + element[5:])
# working on XLS file
yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
gray   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
top_border_thin              = Border(top=Side(style='thin'))
bottom_border_thin           = Border(bottom=Side(style='thin'))
left_border_thin             = Border(left=Side(style='thin'))
right_border_thin            = Border(right=Side(style='thin'))
top_bottom_border_thin       = Border(top=Side(style='thin'), bottom=Side(style='thin'))
top_bottom_left_border_thin  = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
top_bottom_right_border_thin = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
all_border_thin              = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
top_border_thick              = Border(top=Side(style='thick'))
bottom_border_thick           = Border(bottom=Side(style='thick'))
left_border_thick             = Border(left=Side(style='thick'))
right_border_thick            = Border(right=Side(style='thick'))
top_bottom_border_thick       = Border(top=Side(style='thick'), bottom=Side(style='thick'))
top_bottom_left_border_thick  = Border(top=Side(style='thick'), bottom=Side(style='thick'), left=Side(style='thick'))
top_bottom_right_border_thick = Border(top=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thick'))
all_border_thick              = Border(top=Side(style='thick'), bottom=Side(style='thick'), left=Side(style='thick'), right=Side(style='thick'))
wb = Workbook()
sheet = wb.active
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.column_dimensions['A'].width = 30
sheet.column_dimensions['B'].width = 25
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 15
calibri11Bold = Font(size=11, bold=True)
calibri16Bold = Font(size=16, bold=True)
sheet["A1"] = "Obdobje oglaševanja"
sheet["B1"] = "{} {}".format(months[int(prev_month_month)], prev_month_year)
sheet["A3"] = "Kliki"
sheet["A4"] = "Ogledi"
sheet["A5"] = "CPC"
sheet["A6"] = "Poraba"
sheet["A8"] = "Stroški upravljanja"
sheet["A1"].font = calibri11Bold
sheet["B1"].font = calibri11Bold
sheet["B1"].alignment = Alignment(horizontal="right")
sheet["A3"].font = calibri11Bold
sheet["A4"].font = calibri11Bold
sheet["A5"].font = calibri11Bold
sheet["A6"].font = calibri11Bold
counter = 11
clicks       = 0
impressions  = 0
cpc          = 0
cpc_count    = 0
cost         = 0
expense      = 0
for stat in stats:
    if stat[0] != "nolabel":
        sheet["A" + str(counter)]     = "Kategorija " + stat[0]
        sheet["A" + str(counter)].font = calibri11Bold
        counter += 1
    sheet["A" + str(counter)] = "Oglasna skupina"
    sheet["B" + str(counter)] = "Kampanja"
    sheet["C" + str(counter)] = "Kliki"
    sheet["D" + str(counter)] = "Ogledi"
    sheet["E" + str(counter)] = "CPC"
    sheet["F" + str(counter)] = "Poraba"
    sheet["A" + str(counter)].font = calibri11Bold
    sheet["B" + str(counter)].font = calibri11Bold
    sheet["C" + str(counter)].font = calibri11Bold
    sheet["D" + str(counter)].font = calibri11Bold
    sheet["E" + str(counter)].font = calibri11Bold
    sheet["F" + str(counter)].font = calibri11Bold
    sheet["A" + str(counter)].fill = gray
    sheet["B" + str(counter)].fill = gray
    sheet["C" + str(counter)].fill = gray
    sheet["D" + str(counter)].fill = gray
    sheet["E" + str(counter)].fill = gray
    sheet["F" + str(counter)].fill = gray
    sheet["C" + str(counter)].alignment = Alignment(horizontal="right")
    sheet["D" + str(counter)].alignment = Alignment(horizontal="right")
    sheet["E" + str(counter)].alignment = Alignment(horizontal="right")
    sheet["F" + str(counter)].alignment = Alignment(horizontal="right")
    for column in ["A", "B", "C", "D", "E", "F"]:
        if column == "A": sheet[column + str(counter)].border = top_bottom_left_border_thick
        elif column == "F": sheet[column + str(counter)].border = top_bottom_right_border_thick
        else: sheet[column + str(counter)].border = top_bottom_border_thick
    start = counter + 1
    adgroups = reversed(sorted((stat[1])))
    for adgroup in adgroups:
        counter += 1
        sheet["A" + str(counter)] = adgroup[4]
        sheet["B" + str(counter)] = adgroup[5]
        sheet["C" + str(counter)] = adgroup[0]
        sheet["D" + str(counter)] = adgroup[1]
        sheet["E" + str(counter)] = adgroup[2]
        sheet["F" + str(counter)] = adgroup[3]
        if adgroup[0] != 0: sheet["C" + str(counter)].number_format = '#,###'
        if adgroup[1] != 0: sheet["D" + str(counter)].number_format = '#,###'
        sheet["E" + str(counter)].number_format = '#,##0.00 €'
        sheet["F" + str(counter)].number_format = '#,##0.00 €'
    for column in ["A", "B", "C", "D", "E", "F"]: sheet[column + str(counter)].border = bottom_border_thin
    sum_clicks      = 0
    sum_impressions = 0
    sum_cpc         = 0
    sum_cost        = 0
    for i in range(start, counter + 1):
        sum_clicks      += sheet["C" + str(i)].value
        sum_impressions += sheet["D" + str(i)].value
        sum_cpc         += sheet["E" + str(i)].value
        sum_cost        += sheet["F" + str(i)].value
    clicks      += sum_clicks
    impressions += sum_impressions
    cpc         += sum_cpc
    cost        += sum_cost
    counter += 1
    sheet["B" + str(counter)] = "Poraba"
    sheet["C" + str(counter)] = sum_clicks
    sheet["D" + str(counter)] = sum_impressions
    sheet["E" + str(counter)] = sum_cpc / (counter - start)
    sheet["F" + str(counter)] = sum_cost
    sheet["C" + str(counter)].number_format = '#,###'
    sheet["D" + str(counter)].number_format = '#,###'
    sheet["E" + str(counter)].number_format = '#,##0.00 €'
    sheet["F" + str(counter)].number_format = '#,##0.00 €'
    sheet["B" + str(counter)].font = calibri11Bold
    sheet["C" + str(counter)].font = calibri11Bold
    sheet["D" + str(counter)].font = calibri11Bold
    sheet["E" + str(counter)].font = calibri11Bold
    sheet["F" + str(counter)].font = calibri11Bold
    cpc_count += counter - start
    counter += 1
    sheet["B" + str(counter)] = "Stroški upravljanja"
    sheet["F" + str(counter)] = sheet["F" + str(counter - 1)].value * 0.2
    expense += sheet["F" + str(counter - 1)].value * 0.2
    sheet["F" + str(counter)].number_format = '#,##0.00 €'
    sheet["B" + str(counter)].font = calibri11Bold
    sheet["F" + str(counter)].font = calibri11Bold
    for column in ["B", "C", "D", "E", "F"]: sheet[column + str(counter)].border = bottom_border_thick
    counter += 1
    sheet["B" + str(counter)] = "SKUPAJ"
    sheet["F" + str(counter)] = sheet["F" + str(counter - 2)].value * 1.07
    sheet["F" + str(counter)].number_format = '#,##0.00 €'
    sheet["F" + str(counter)].border = all_border_thin
    sheet["B" + str(counter)].font = calibri11Bold
    sheet["F" + str(counter)].font = calibri11Bold
    counter += 1
sheet["B3"] = clicks
sheet["B4"] = impressions
sheet["B5"] = cpc / cpc_count
sheet["B6"] = cost
sheet["B8"] = expense
sheet["B3"].number_format = '#,###'
sheet["B4"].number_format = '#,###'
sheet["B5"].number_format = '#,##0.00 €'
sheet["B6"].number_format = '#,##0.00 €'
sheet["B8"].number_format = '#,##0.00 €'
sheet["B3"].font = calibri11Bold
sheet["B4"].font = calibri11Bold
sheet["B5"].font = calibri11Bold
sheet["B6"].font = calibri11Bold
sheet["A8"].font = calibri16Bold
sheet["B8"].font = calibri16Bold
sheet["A8"].fill = yellow
sheet["B8"].fill = yellow
sheet["A8"].border = top_bottom_left_border_thick
sheet["B8"].border = top_bottom_right_border_thick
while True:
    try:
        wb.save(path + "Avtenta_Specifikacija_po_kategorijah_{}_{}.xlsx".format(prev_month_year, prev_month_month))
        break
    except PermissionError: input("\nPlease close Avtenta_Specifikacija_po_kategorijah_{}_{}.xlsx and press any key. ".format(prev_month_year, prev_month_month))
print("\nArticle info saved as Avtenta_Specifikacija_po_kategorijah_{}_{}.xlsx".format(prev_month_year, prev_month_month))