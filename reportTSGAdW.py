from os import getcwd
from datetime import date
from csv import reader, writer, register_dialect, QUOTE_NONE
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Alignment

def getCSVdata(file, dlm=",", enc="UTF-8"):
    """
    Gets data from a CSV file and returns it as a 2D list.
    :param file: str; CSV file to process
    :param dlm: str; delimiter, default ,
    :param enc: str; encoding, default UTF-8
    :return: list; CSV data in a 2D list
    """
    with open(file, encoding=enc) as stats:
        return list(reader(stats, delimiter=dlm))

# creates path
path = getcwd().replace("\\", "\\\\") + "\\\\"
# date variables
today = date.today()
prev_month_year = today.year
prev_month_month = today.month - 1
if prev_month_month == 0:
    prev_month_year -= 1
    prev_month_month = 12
prev_month_month = str(prev_month_month)
months = { 1: "Januar",
           2: "Februar",
           3: "Marec",
           4: "April",
           5: "Maj",
           6: "Junij",
           7: "Julij",
           8: "Avgust",
           9: "September",
          10: "Oktober",
          11: "November",
          12: "December"}
if len(prev_month_month) == 1: prev_month_month = "0" + prev_month_month
prev_month_year = str(prev_month_year)
# importing CSV data
register_dialect('myDialect', delimiter=';', quoting=QUOTE_NONE)
GAdW_prev = getCSVdata(path.replace("\\\\", "/") + "GAdW_Telekom_CSV_{}-{}.csv".format(prev_month_year, prev_month_month), ";")
stats = [["Mobi", []], ["Mobilni internet", []], ["Mobitel", []], ["Poslovni", []], ["SiOL", []], ["Telekom", []]]
for element in GAdW_prev:
    for stat in sorted(stats):
        if element[0] == stat[0]:
            stat[1].append([float(item) for item in element[1:4]] + element[4:])
# working on XLS file
yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
gray   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
top_border_thin              = Border(top=Side(style='thin'))
bottom_border_thin           = Border(bottom=Side(style='thin'))
left_border_thin             = Border(left=Side(style='thin'))
right_border_thin            = Border(right=Side(style='thin'))
top_bottom_border_thin       = Border(top=Side(style='thin'), bottom=Side(style='thin'))
top_bottom_left_border_thin  = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
top_bottom_right_border_thin = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
all_border_thin              = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
top_border_thick              = Border(top=Side(style='thick'))
bottom_border_thick           = Border(bottom=Side(style='thick'))
left_border_thick             = Border(left=Side(style='thick'))
right_border_thick            = Border(right=Side(style='thick'))
top_bottom_border_thick       = Border(top=Side(style='thick'), bottom=Side(style='thick'))
top_bottom_left_border_thick  = Border(top=Side(style='thick'), bottom=Side(style='thick'), left=Side(style='thick'))
top_bottom_right_border_thick = Border(top=Side(style='thick'), bottom=Side(style='thick'), right=Side(style='thick'))
all_border_thick              = Border(top=Side(style='thick'), bottom=Side(style='thick'), left=Side(style='thick'), right=Side(style='thick'))
wb = Workbook()
sheet = wb.active
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.column_dimensions['A'].width = 50
sheet.column_dimensions['B'].width = 25
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15
calibri11Bold = Font(size=11, bold=True)
calibri16Bold = Font(size=16, bold=True)
sheet["A1"] = "Obdobje oglaševanja"
sheet["B1"] = "{} {}".format(months[int(prev_month_month)], prev_month_year)
sheet["A3"] = "Interakcije (kliki, ogledi)"
sheet["A4"] = "CPC"
sheet["A5"] = "Poraba"
sheet["A7"] = "Stroški upravljanja"
sheet["A1"].font = calibri11Bold
sheet["B1"].font = calibri11Bold
sheet["B1"].alignment = Alignment(horizontal="right")
sheet["A3"].font = calibri11Bold
sheet["A4"].font = calibri11Bold
sheet["A5"].font = calibri11Bold
counter = 10
interactions = 0
cpc          = 0
cpc_count    = 0
cost         = 0
expense      = 0
for stat in stats:
    sheet["A" + str(counter)]     = "Kategorija " + stat[0]
    sheet["A" + str(counter)].font = calibri11Bold
    counter += 1
    sheet["A" + str(counter)] = "Oglasna skupina"
    sheet["B" + str(counter)] = "Kampanja"
    if stat[0] == "Telekom": sheet["C" + str(counter)] = "Kliki/ogledi"
    else: sheet["C" + str(counter)] = "Kliki"
    sheet["D" + str(counter)] = "CPC"
    sheet["E" + str(counter)] = "Poraba"
    sheet["A" + str(counter)].font = calibri11Bold
    sheet["B" + str(counter)].font = calibri11Bold
    sheet["C" + str(counter)].font = calibri11Bold
    sheet["D" + str(counter)].font = calibri11Bold
    sheet["E" + str(counter)].font = calibri11Bold
    sheet["A" + str(counter)].fill = gray
    sheet["B" + str(counter)].fill = gray
    sheet["C" + str(counter)].fill = gray
    sheet["D" + str(counter)].fill = gray
    sheet["E" + str(counter)].fill = gray
    sheet["C" + str(counter)].alignment = Alignment(horizontal="right")
    sheet["D" + str(counter)].alignment = Alignment(horizontal="right")
    sheet["E" + str(counter)].alignment = Alignment(horizontal="right")
    for column in ["A", "B", "C", "D", "E"]:
        if column == "A": sheet[column + str(counter)].border = top_bottom_left_border_thick
        elif column == "E": sheet[column + str(counter)].border = top_bottom_right_border_thick
        else: sheet[column + str(counter)].border = top_bottom_border_thick
    start = counter + 1
    adgroups = reversed(sorted((stat[1])))
    for adgroup in adgroups:
        counter += 1
        sheet["A" + str(counter)] = adgroup[3]
        sheet["B" + str(counter)] = adgroup[4]
        sheet["C" + str(counter)] = adgroup[0]
        sheet["D" + str(counter)] = adgroup[1]
        sheet["E" + str(counter)] = adgroup[2]
        sheet["B" + str(counter)].number_format = '#,###'
        if adgroup[0] != 0: sheet["C" + str(counter)].number_format = '#,###'
        sheet["D" + str(counter)].number_format = '#,##0.00 €'
        sheet["E" + str(counter)].number_format = '#,##0.00 €'
    for column in ["A", "B", "C", "D", "E"]: sheet[column + str(counter)].border = bottom_border_thin
    sum_clicks = 0
    sum_cpc    = 0
    sum_cost   = 0
    for i in range(start, counter + 1):
        sum_clicks += sheet["C" + str(i)].value
        sum_cpc    += sheet["D" + str(i)].value
        sum_cost   += sheet["E" + str(i)].value
    interactions += sum_clicks
    cpc          += sum_cpc
    cost         += sum_cost
    counter += 1
    sheet["B" + str(counter)] = "Poraba"
    sheet["C" + str(counter)] = sum_clicks
    sheet["D" + str(counter)] = sum_cpc / (counter - start)
    sheet["E" + str(counter)] = sum_cost
    sheet["C" + str(counter)].number_format = '#,###'
    sheet["D" + str(counter)].number_format = '#,##0.00 €'
    sheet["E" + str(counter)].number_format = '#,##0.00 €'
    sheet["B" + str(counter)].font = calibri11Bold
    sheet["C" + str(counter)].font = calibri11Bold
    sheet["D" + str(counter)].font = calibri11Bold
    sheet["E" + str(counter)].font = calibri11Bold
    cpc_count += counter - start
    counter += 1
    sheet["B" + str(counter)] = "Stroški upravljanja"
    sheet["E" + str(counter)] = sheet["E" + str(counter - 1)].value * 0.07
    expense += sheet["E" + str(counter - 1)].value * 0.07
    sheet["E" + str(counter)].number_format = '#,##0.00 €'
    sheet["B" + str(counter)].font = calibri11Bold
    sheet["E" + str(counter)].font = calibri11Bold
    for column in ["B", "C", "D", "E"]: sheet[column + str(counter)].border = bottom_border_thick
    counter += 1
    sheet["B" + str(counter)] = "SKUPAJ"
    sheet["E" + str(counter)] = sheet["E" + str(counter - 2)].value * 1.07
    sheet["E" + str(counter)].number_format = '#,##0.00 €'
    sheet["E" + str(counter)].border = all_border_thin
    sheet["B" + str(counter)].font = calibri11Bold
    sheet["E" + str(counter)].font = calibri11Bold
    counter += 1
sheet["B3"] = interactions
sheet["B4"] = cpc / cpc_count
sheet["B5"] = cost
sheet["B7"] = expense
sheet["B3"].number_format = '#,###'
sheet["B4"].number_format = '#,##0.00 €'
sheet["B5"].number_format = '#,##0.00 €'
sheet["B7"].number_format = '#,##0.00 €'
sheet["B3"].font = calibri11Bold
sheet["B4"].font = calibri11Bold
sheet["B5"].font = calibri11Bold
sheet["A7"].font = calibri16Bold
sheet["B7"].font = calibri16Bold
sheet["A7"].fill = yellow
sheet["B7"].fill = yellow
sheet["A7"].border = top_bottom_left_border_thick
sheet["B7"].border = top_bottom_right_border_thick
while True:
    try:
        wb.save(path + "TS_Specifikacija_po_kategorijah_{}_{}.xlsx".format(prev_month_year, prev_month_month))
        break
    except PermissionError: input("\nPlease close TS_Specifikacija_po_kategorijah_{}_{}.xlsx and press any key. ".format(prev_month_year, prev_month_month))
print("\nArticle info saved as TS_Specifikacija_po_kategorijah_{}_{}.xlsx".format(prev_month_year, prev_month_month))